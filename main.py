import datetime
import openpyxl
from openpyxl.utils import column_index_from_string
import openpyxl.styles as style


class Managers:

    def __init__(self, file_channels: str, col_managers='E', col_contracts='G'):
        self.file_channels = file_channels
        self._col_managers = col_managers
        self._col_contracts = col_contracts

    @property
    def contracts(self):
        return self.__class__._parse_file(self)

    def _parse_file(self):
        """
        Загружает из файла 'Договора.xlsx' список номеров договоров с привязкой к куратору
        :return:
        """
        wb = openpyxl.load_workbook(filename=self.file_channels, read_only=True)
        sheet = wb.active
        manager_c_num = column_index_from_string(self._col_managers)
        contract_c_num = column_index_from_string(self._col_contracts)
        result = {}
        # создаем словарь менеджеров
        for r_num in range(2, sheet.max_row):  # цикл по строкам в таблице
            key = sheet.cell(r_num, manager_c_num).value
            if key:
                result[str(key).strip()] = []
        # заполняем словарь менеджеров номерами договоров
        for r_num in range(2, sheet.max_row):  # цикл по строкам в таблице
            key = sheet.cell(r_num, manager_c_num).value
            if key:
                value = sheet.cell(r_num, contract_c_num).value
                result[key].append(str(value).strip())
        wb.close()
        return result


class Reports:

    def __init__(self, file_report: str, managers: dict, col_contracts=5):
        self.__file_report = file_report
        self.__managers = managers
        self.col_contracts = col_contracts
        self.list_of_head = ['Местонахождения Точки № АЗС',
                             'Дата и время начала перерыва',
                             'Дата и время окончания перерыва',
                             'Длительность перерыва (часов)',
                             'Причина перерыва',
                             'Принятые меры № ТТ',
                             '№ Заявки',
                             '№ TT']
        self._tmp_report = self._make_tmp_report()

    def _get_source_report(self):
        """
        Загружает из файла 'КПД.xlsx' всю исходную информацию по простоям каналов.
        Исходный файл 'КПД.xlsx' формируется с помощью экспорта из CMDB заявок за предыдущий месяц.
        В CMDB используется представление !отчет_КПД (Table)
        :return:
        Возвращает страницу из файла в виде рабочей книги Excel.
        """
        return openpyxl.load_workbook(filename=self.__file_report)

    @staticmethod
    def _contract_in_cell(contract_list, data_cell):
        """
        Проверяет наличие номера договора в ячейке
        :param contract_list:
        :param data_cell:
        :return: возвращает True в случае успеха, иначе False
        """
        result = False
        for cont in contract_list:
            if cont in str(data_cell):
                result = True
                break
        return result

    def _make_tmp_report(self):
        """
        Создается промежуточный вариант исходного отчета путем добавления
        к каждой строке отчета ФИО менеджера
        :return:
        """
        tmp_report_wb = self._get_source_report()
        tmp_report_sheet = tmp_report_wb.active
        for r_num in range(2, tmp_report_sheet.max_row):
            contract_data_cell = tmp_report_sheet.cell(r_num, self.col_contracts).value
            for manager in self.__managers:
                if self._contract_in_cell(self.__managers[manager], contract_data_cell):
                    # добавить в последний столбец имя манагера
                    # и выйти из цикла по манагерам
                    _ = tmp_report_sheet.cell(r_num, 9, value=manager)
                    continue
                else:
                    # перейти к следующему манагеру
                    continue
        return tmp_report_wb

    @staticmethod
    def _format_report_title(ws_report):
        """
        Форматирование отчета
        :param ws_report:
        :return:
        """
        # форматирование заголовков
        for char in "ABCDEFGH":
            ws_report.column_dimensions[char].width = 30
            ws_report[char + '1'].font = style.Font(bold=True, size=13)
            ws_report[char + '1'].alignment = style.Alignment(wrap_text=True)
        # форматирование столбца с D
        for index in range(2, ws_report.max_row + 1):
            _cell = ws_report[f'D{index}']
            _cell.number_format = '[h]:mm:ss'

    def _make_report(self, manager):
        """
        Создается отчет по каналам конкретного менеджера
        :param manager:
        :return: возвращается объект Workbook или None если по данному
        менеджеру не найдено ни одной строки
        """
        current_index_row = 1
        manager_in_report = False
        tmp_report_sheet = self._tmp_report.active
        report_wb = openpyxl.Workbook()
        report_ws = report_wb.active
        report_ws.append(self.list_of_head)
        for r_num in range(2, tmp_report_sheet.max_row):
            if tmp_report_sheet.cell(r_num, 9).value == manager:
                manager_in_report = True
                tmp_row = []
                for c_num in range(1, 9):
                    tmp_row.append(tmp_report_sheet.cell(r_num, c_num).value)
                report_ws.append(tmp_row)
                current_index_row += 1
                # вставка формулы в столбец D
                report_ws[f"D{current_index_row}"] = f"=C{current_index_row}-B{current_index_row}"
        self._format_report_title(report_ws)
        report_wb.close()
        if manager_in_report:
            return report_wb
        else:
            return None

    def _make_reports(self):
        """
        Создается словарь в котором ключи ФИО менеджеров,
        а значение объект отчета данного менеджера
        :return: возвращается словарь с объектами Workbook
        """
        dict_of_reports = {}
        for manager in self.__managers:
            report_wb = self._make_report(manager)
            if report_wb:
                dict_of_reports[manager] = report_wb
        return dict_of_reports

    @staticmethod
    def _get_report_month():
        today = datetime.date.today()
        first = today.replace(day=1)
        last = first - datetime.timedelta(days=1)
        return last.strftime("%Y.%m")

    def write_file_report(self):
        """
        Созданные отчеты записываются в отдельные файлы
        :return: None
        """
        dict_of_reports = self._make_reports()
        month_of_report = self._get_report_month()
        for manager in dict_of_reports:
            dict_of_reports[manager].save(f"{month_of_report} {manager} отчет по простоям.xlsx")


if __name__ == '__main__':
    report = 'files\\КПД.xlsx'
    contract = 'files\\Договора.xlsx'
    mg = Managers(contract)
    rep = Reports(report, mg.contracts)
    rep.write_file_report()
